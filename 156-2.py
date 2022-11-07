from openpyxl import load_workbook

wb = load_workbook("dodgers.xlsx")
result = []

ws = wb.worksheets[0]
for row in ws.iter_rows():
    # result.append([cell.value for cell in row])
    tmp_list = []
    for cell in row:
        tmp_list.append(cell.value)
    result.append(tmp_list)

print(result)

sum = 0
for r in result[1:]:  # [0]是標題列，不計算
    sum += r[11]  # 將[11](HR)的值加總
print(sum)
