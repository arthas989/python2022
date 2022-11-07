from openpyxl import load_workbook

wb = load_workbook("dodgers.xlsx")
result = []

ws = wb.worksheets[0]  # 第一個worksheet
for row in ws.iter_rows():
    # list comprehension
    # result.append([cell.value for cell in row])
    '''
    第一次改時改錯，全部append成一個list了，要有多個list of list才對
    for cell in row:
        result.append(cell.value)
    '''
    tmp_list = []
    for cell in row:
        tmp_list.append(cell.value)
    result.append(tmp_list)
print(result)

sum = 0
for r in result[1:]:
    print(r)
    # sum += int(r[11])

print(f"The total number of homeruns for Dodgers was {sum}.")
